import File1
from File1 import selection, stock
import datetime
from openpyxl import Workbook, load_workbook
import os

# Define the items and their prices
items = {'Coke': 1.5, 'Chips': 1.25, 'Candy': 0.75}

# Check if the sales.xlsx file exists
if os.path.exists('sales.xlsx'):
    wb = load_workbook('sales.xlsx')
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Time', 'Item', 'Quantity', 'Price', 'Total'])
    wb.save('sales.xlsx')

# Collect the data
while True:
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    item_name = selection
    if item_name == 'exit':
        break
    if item_name not in items:
        print("Invalid item")
        continue
    if stock[item_name] == 0:
        print("Out of stock")
        break
    item_price = items[item_name]
    initial_stock_level = stock[item_name]
    stock[item_name] -= 1
    current_stock_level = stock[item_name]
    quantity_sold = initial_stock_level - current_stock_level
    total_sales = quantity_sold * item_price

    # Add the data to the Excel sheet
    ws.append([formatted_time, item_name, quantity_sold, item_price, total_sales])
    wb.save('sales.xlsx')

    # Print some values for debugging purposes
    print(f"Item: {item_name}")
    print(f"Stock: {stock[item_name]}")
    print(f"Total sales: {total_sales}")

    # Prompt the user to continue or exit the program
    response = input("Enter 'exit' to end the program, or press enter to continue: ")
    if response.lower() == 'exit':
        break

print("Sales data collected and stored in Excel document.")
