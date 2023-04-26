import datetime
from openpyxl import Workbook, load_workbook
import os

# Define the items and their prices
items = {'Coke': 1.5, 'Chips': 1.25, 'Candy': 0.75}

# Collect the data
while True:
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    item_name = input("Enter the item sold (type 'exit' to quit): ")
    if item_name == 'exit':
        break
    if item_name not in items:
        print("Invalid item")
        continue
    item_price = items[item_name]
    item_quantity = int(input("Enter the number of {0} sold: ".format(item_name)))
    total_sales = item_quantity * item_price

    # Add the data to an Excel worksheet
    if os.path.exists('sales.xlsx'):
        wb = load_workbook('sales.xlsx')
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Time', 'Item', 'Quantity', 'Price', 'Total'])
    ws.append([formatted_time, item_name, item_quantity, item_price, total_sales])
    wb.save('sales.xlsx')

print("Sales data collected and stored in Excel document.")

 